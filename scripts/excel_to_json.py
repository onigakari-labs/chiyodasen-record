"""
excel_to_json.py
records/chiyodasen_record.xlsx を読み込んで、
docs/data/record.json として書き出す。

HTMLダッシュボードはこのJSONをfetchして表示する。
"""
import json
import os
from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook

ROOT = os.path.join(os.path.dirname(__file__), "..")
EXCEL_PATH = os.path.join(ROOT, "records", "chiyodasen_record.xlsx")
JSON_PATH = os.path.join(ROOT, "docs", "data", "record.json")

JST = timezone(timedelta(hours=9))


def normalize_date(value):
    """日付をYYYY/MM/DD形式の文字列に正規化"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    s = str(value).strip()
    return s


def normalize_int(value):
    """数値を int に。空欄は None"""
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def main():
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel が見つかりません: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws_record = wb["戦績"]
    ws_races = wb["レース別"]

    # ===== 戦績シート =====
    weeks = {}
    for row in ws_record.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 日付なし行はスキップ
            continue
        date = normalize_date(row[0])
        weeks[date] = {
            "date": date,
            "course": row[1] or "",
            "g1": row[2] or "",
            "axisSingle": normalize_int(row[3]),
            "axisDouble": normalize_int(row[4]),
            "broad": normalize_int(row[5]),
            "perfectHit": (row[6] == "○"),
            "aiEval": row[7] or "",
            "memo": row[8] or "",
            "shareUrl": row[9] or "",
            "races": [],
        }

    # ===== レース別シート =====
    for row in ws_races.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        date = normalize_date(row[0])
        if date not in weeks:
            print(f"WARN: {date} は戦績シートに存在しません、スキップ")
            continue
        race = {
            "r": normalize_int(row[1]),
            "course": row[2] or "",
            "name": row[3] or "",
            "marks": {
                "◎": normalize_int(row[4]),
                "○": normalize_int(row[5]),
                "▲": normalize_int(row[6]),
                "△": normalize_int(row[7]),
                "×": normalize_int(row[8]),
            },
            "first": normalize_int(row[9]),
            "horseName": row[10] or "",
            "axisHit": (row[11] == "✓"),
        }
        weeks[date]["races"].append(race)

    # ===== 集計 =====
    weeks_list = sorted(weeks.values(), key=lambda w: w["date"], reverse=True)
    total_weeks = len(weeks_list)
    total_races = sum(len(w["races"]) for w in weeks_list)
    axis_hit_races = sum(1 for w in weeks_list for r in w["races"] if r["axisHit"])

    summary = {
        "totalWeeks": total_weeks,
        "totalRaces": total_races,
        "axisSingleHits": sum((w["axisSingle"] or 0) for w in weeks_list),
        "axisDoubleHits": sum((w["axisDouble"] or 0) for w in weeks_list),
        "broadHits": sum((w["broad"] or 0) for w in weeks_list),
        "perfectHitWeeks": sum(1 for w in weeks_list if w["perfectHit"]),
        "axisHitRate": (axis_hit_races / total_races) if total_races > 0 else 0.0,
        "axisHitRaces": axis_hit_races,
    }

    payload = {
        "generatedAt": datetime.now(JST).isoformat(),
        "summary": summary,
        "weeks": weeks_list,
    }

    # ===== 書き出し =====
    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    abs_path = os.path.abspath(JSON_PATH)
    size = os.path.getsize(JSON_PATH)
    print(f"OK: {abs_path}")
    print(f"   {size} bytes")
    print(f"   累計: {total_weeks}週 / {total_races}レース")
    print(f"   ◎的中率: {summary['axisHitRate']:.1%}")


if __name__ == "__main__":
    main()