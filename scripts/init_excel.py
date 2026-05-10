"""
init_excel.py
chiyodasen_record.xlsx の初期テンプレートを生成し、
Phase 3-1 で取得した 20260510 のデータを投入する。

シート構成:
  - 戦績: 1日のWIN5サマリ（軸戦略ごとの的中数、メモ等）
  - レース別: 1レース単位の印・結果・的中フラグ
  - 集計: 累計サマリ（自動計算式）
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "records", "chiyodasen_record.xlsx")

# 千代田線カラー
ACCENT_GOLD = "C8A96A"
DARK_BG = "1A1A1A"
LIGHT_BG = "F5F2EA"
HEADER_BG = "2C2C2A"
HIT_BG = "FFF8DC"  # 的中時の薄黄
MISS_BG = "F0F0F0"

thin = Side(border_style="thin", color="888888")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def setup_header(ws, headers, widths):
    """ヘッダ行を整形"""
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def add_data_row(ws, row_idx, values, hit_columns=None):
    """データ行を追加。hit_columns: 的中セル(列番号list、1-index)はゴールド系で塗る"""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="游ゴシック", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        if hit_columns and col_idx in hit_columns:
            cell.fill = PatternFill("solid", fgColor=HIT_BG)
            cell.font = Font(name="游ゴシック", size=10, bold=True)


def main():
    wb = Workbook()

    # ===== シート1: 戦績 =====
    ws1 = wb.active
    ws1.title = "戦績"
    headers1 = [
        "日付", "開催", "G1/重賞",
        "軸1頭(/5)", "軸2頭BOX(/5)", "総流し(/5)",
        "完全的中", "AI評価", "メモ", "公開URL"
    ]
    widths1 = [12, 18, 14, 12, 14, 12, 11, 12, 30, 50]
    setup_header(ws1, headers1, widths1)

    # 20260510 のデータ
    add_data_row(ws1, 2, [
        "2026/05/10",
        "京都/東京/新潟",
        "NHKマイルC(G1)",
        2,
        3,
        3,
        "×",
        "○",
        "NHKマイル ◎17 ロデオドライブ的中。橘S◎4も的中",
        "https://chiyodasen.onrender.com/share/9h4FSQk8JBvl9zvKP1vKrz0A",
    ], hit_columns=[4, 5, 6])

    ws1.row_dimensions[1].height = 28

    # ===== シート2: レース別 =====
    ws2 = wb.create_sheet("レース別")
    headers2 = [
        "日付", "R", "コース", "レース名",
        "◎", "○", "▲", "△", "×",
        "1着", "馬名", "◎的中"
    ]
    widths2 = [12, 6, 10, 18, 6, 6, 6, 6, 6, 8, 18, 8]
    setup_header(ws2, headers2, widths2)

    # 5/10 の5レース
    races = [
        ("2026/05/10", 1, "京都10R", "橘S", 4, 5, 1, 6, 10, 4, "タガノアラリア", "✓"),
        ("2026/05/10", 2, "東京10R", "メトロポリタ", 10, 2, 12, 8, 5, 12, "ウエストナウ", "×"),
        ("2026/05/10", 3, "新潟11R", "谷川岳S", 11, 12, 9, 5, 6, 7, "ランフォーヴァウ", "×"),
        ("2026/05/10", 4, "京都11R", "平城京S", 10, 4, 16, 5, 14, 4, "テスティモーネ", "×"),
        ("2026/05/10", 5, "東京11R", "NHKマイル", 17, 10, 7, 16, 4, 17, "ロデオドライブ", "✓"),
    ]

    for idx, race in enumerate(races, start=2):
        is_hit = race[-1] == "✓"
        # 馬番(◎)と1着が一致したら的中、その行をゴールド系
        hit_cols = [12] if is_hit else None
        add_data_row(ws2, idx, race, hit_columns=hit_cols)

    ws2.row_dimensions[1].height = 28

    # ===== シート3: 集計（自動計算式付き） =====
    ws3 = wb.create_sheet("集計")
    ws3["A1"] = "千代田線 WIN5 累計成績"
    ws3["A1"].font = Font(name="游ゴシック", size=14, bold=True)
    ws3.merge_cells("A1:C1")

    ws3["A3"] = "累計予想週数"
    ws3["B3"] = "=COUNTA(戦績!A2:A1000)"
    ws3["A4"] = "軸1頭 累計的中数"
    ws3["B4"] = '=SUMIFS(戦績!D:D,戦績!D:D,">0")'
    ws3["A5"] = "軸2頭BOX 累計的中数"
    ws3["B5"] = '=SUMIFS(戦績!E:E,戦績!E:E,">0")'
    ws3["A6"] = "総流し 累計的中数"
    ws3["B6"] = '=SUMIFS(戦績!F:F,戦績!F:F,">0")'
    ws3["A7"] = "完全的中回数"
    ws3["B7"] = '=COUNTIF(戦績!G:G,"○")'

    ws3["A9"] = "レース別 ◎的中率"
    ws3["A9"].font = Font(bold=True)
    ws3["A10"] = "総レース数"
    ws3["B10"] = "=COUNTA(レース別!A2:A10000)"
    ws3["A11"] = "◎的中数"
    ws3["B11"] = '=COUNTIF(レース別!L:L,"✓")'
    ws3["A12"] = "◎的中率"
    ws3["B12"] = "=B11/B10"
    ws3["B12"].number_format = "0.0%"

    for col, w in [("A", 26), ("B", 16)]:
        ws3.column_dimensions[col].width = w

    # ===== 保存 =====
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    abs_path = os.path.abspath(OUTPUT_PATH)
    size = os.path.getsize(OUTPUT_PATH)
    print(f"OK: {abs_path}")
    print(f"   {size} bytes")


if __name__ == "__main__":
    main()